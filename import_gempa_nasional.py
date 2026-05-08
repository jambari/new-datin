"""
import_gempa_nasional.py — Import all non-Papua sheets from the national catalog.

Skips sheets already imported (Papua, Papua Barat, Papua Tengah, Papua Pegunungan)
and empty sheets. Deduplicates against existing DB rows by (tanggal, lat, lng, mag).

Run from project root:
    source venv/bin/activate && python import_gempa_nasional.py
"""
import os, sys, django, datetime, re
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "datin_project.settings")
sys.path.insert(0, os.path.dirname(__file__))
django.setup()

import openpyxl
from repository.models import GempaMemusak

XLSX = "/mnt/d/data/Katalog Gempa Perwilayah 1821- 2024 update.xlsx"

# Sheet name → canonical province name
SHEET_TO_PROVINCE = {
    'Aceh':                 'Aceh',
    'Sumut':                'Sumatera Utara',
    'SumSel':               'Sumatera Selatan',
    'Sumbar':               'Sumatera Barat',
    'Bengkulu':             'Bengkulu',
    'Jambi':                'Jambi',
    'Lampung':              'Lampung',
    'Banten':               'Banten',
    'DKI Jakarta':          'DKI Jakarta',
    'Jabar':                'Jawa Barat',
    'Jateng':               'Jawa Tengah',
    'Jatim':                'Jawa Timur',
    'DIY':                  'DI Yogyakarta',
    'Bali':                 'Bali',
    'Nusa Tenggara Timur':  'Nusa Tenggara Timur',
    'Nusa Tenggara Barat':  'Nusa Tenggara Barat',
    'KalBar':               'Kalimantan Barat',
    'Kaltim':               'Kalimantan Timur',
    'KalSel':               'Kalimantan Selatan',
    'KalTara':              'Kalimantan Utara',
    'Gorontalo':            'Gorontalo',
    'Sulbar':               'Sulawesi Barat',
    'SulTengah':            'Sulawesi Tengah',
    'Sulut':                'Sulawesi Utara',
    'SulTenggara':          'Sulawesi Tenggara',
    'Sulsel':               'Sulawesi Selatan',
    'Maluku-Utara':         'Maluku Utara',
    'Maluku':               'Maluku',
}

ID_MONTHS = {
    "januari": 1,  "jan": 1,
    "februari": 2, "feb": 2,
    "maret": 3,    "mar": 3,
    "april": 4,    "apr": 4,
    "mei": 5,
    "juni": 6,     "jun": 6,
    "juli": 7,     "jul": 7,
    "agustus": 8,  "agst": 8, "ags": 8, "agus": 8,
    "september": 9,"sept": 9, "sep": 9,
    "oktober": 10, "okt": 10,
    "november": 11,"nov": 11,
    "desember": 12,"des": 12,
}


def parse_date(val):
    if val is None:
        return None, ''
    if isinstance(val, datetime.datetime):
        return val.date(), val.strftime("%-d %B %Y")
    if isinstance(val, datetime.date):
        return val, val.strftime("%-d %B %Y")
    text = re.sub(r'\s+', ' ', str(val).strip().split('\n')[0])
    parts = text.split()
    if len(parts) >= 3:
        try:
            day   = int(parts[0])
            month = ID_MONTHS.get(parts[1].lower())
            year  = int(parts[2])
            if month:
                return datetime.date(year, month, day), text
        except (ValueError, IndexError):
            pass
    return None, text


def parse_float(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if s in ('-', '', 'WIB'):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def parse_time(val):
    if val is None:
        return None
    if isinstance(val, datetime.time):
        return val
    if isinstance(val, datetime.datetime):
        return val.time()
    s = str(val).strip()
    if s in ('-', '', 'WIB'):
        return None
    try:
        parts = s.split(':')
        if len(parts) >= 2:
            return datetime.time(int(parts[0]), int(parts[1]),
                                 int(float(parts[2])) if len(parts) > 2 else 0)
    except (ValueError, IndexError):
        pass
    return None


def parse_sheet(ws, province):
    all_rows = list(ws.iter_rows(values_only=True))
    data_rows = [r for r in all_rows[7:] if any(c is not None for c in r)]

    blocks = []
    current = None
    for row in data_rows:
        if len(row) < 2:
            continue
        val = row[1]
        if val is not None and isinstance(val, (int, float)) and float(val) == int(float(val)) and val > 0:
            if current:
                blocks.append(current)
            current = {"rows": [row]}
        elif current is not None:
            current["rows"].append(row)
    if current:
        blocks.append(current)

    events = []
    for blk in blocks:
        rows = blk["rows"]
        r0 = rows[0]
        r1 = rows[1] if len(rows) > 1 else [None] * 12

        tanggal, tanggal_txt = parse_date(r0[2] if len(r0) > 2 else None)
        wilayah = str(r1[2]).strip() if len(r1) > 2 and r1[2] else ''
        # Strip "Kab."/"Kota" prefixes for cleaner display
        wilayah_clean = re.sub(r'^(Kab\.|Kota|Kabupaten)\s+', '', wilayah, flags=re.I).strip()

        lokasi_raw = str(r0[8]).strip() if len(r0) > 8 and r0[8] else ''
        lokasi = lokasi_raw if lokasi_raw in ('Darat', 'Laut') else ''

        tsunami = None
        for row in rows:
            cell = str(row[2]).strip().lower() if len(row) > 2 and row[2] else ''
            if cell == 'tsunami':
                tsunami = True
                break
            if cell == 'tidak tsunami':
                tsunami = False
                break

        wil_parts  = [str(r[9]).strip()  for r in rows if len(r) > 9  and r[9]  and str(r[9]).strip()  not in ('-', '')]
        korb_parts = [str(r[10]).strip() for r in rows if len(r) > 10 and r[10] and str(r[10]).strip() not in ('-', '')]

        events.append({
            "tanggal_text":      tanggal_txt,
            "tanggal":           tanggal,
            "wilayah":           wilayah_clean,
            "provinsi":          province,
            "origin_time":       parse_time(r0[3] if len(r0) > 3 else None),
            "latitude":          parse_float(r0[4] if len(r0) > 4 else None),
            "longitude":         parse_float(r0[5] if len(r0) > 5 else None),
            "depth_km":          parse_float(r0[6] if len(r0) > 6 else None),
            "magnitude":         parse_float(r0[7] if len(r0) > 7 else None),
            "lokasi":            lokasi,
            "tsunami":           tsunami,
            "wilayah_merasakan": '\n'.join(wil_parts),
            "korban_kerusakan":  '\n'.join(korb_parts),
            "sumber":            str(r0[11]).strip() if len(r0) > 11 and r0[11] else '',
        })
    return events


def dedup_key(evt):
    lat = round(evt["latitude"],  2) if evt["latitude"]  else None
    lng = round(evt["longitude"], 2) if evt["longitude"] else None
    mag = round(evt["magnitude"], 1) if evt["magnitude"] else None
    return (evt["tanggal"], lat, lng, mag)


# Build dedup set from existing records
existing_keys = set()
for obj in GempaMemusak.objects.values("tanggal", "latitude", "longitude", "magnitude"):
    lat = round(obj["latitude"],  2) if obj["latitude"]  else None
    lng = round(obj["longitude"], 2) if obj["longitude"] else None
    mag = round(obj["magnitude"], 1) if obj["magnitude"] else None
    existing_keys.add((obj["tanggal"], lat, lng, mag))

next_no = (GempaMemusak.objects.order_by('-no').values_list('no', flat=True).first() or 0) + 1

wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
total_created = 0
total_skipped = 0

for sheet_name, province in SHEET_TO_PROVINCE.items():
    if sheet_name not in wb.sheetnames:
        print(f"\n[{sheet_name}] sheet not found — skipping")
        continue
    ws = wb[sheet_name]
    events = parse_sheet(ws, province)
    if not events:
        print(f"\n[{sheet_name}]  0 events — skipping")
        continue

    created = 0
    skipped = 0
    print(f"\n[{sheet_name} → {province}]  {len(events)} events")
    for evt in events:
        key = dedup_key(evt)
        if key in existing_keys:
            skipped += 1
            total_skipped += 1
            continue
        GempaMemusak.objects.create(no=next_no, **evt)
        existing_keys.add(key)
        next_no += 1
        created += 1
        total_created += 1

    print(f"  created={created}  skipped={skipped}")

print(f"\n{'='*50}")
print(f"Total created: {total_created}   Total skipped (dup): {total_skipped}")
print(f"Total GempaMemusak records: {GempaMemusak.objects.count()}")
