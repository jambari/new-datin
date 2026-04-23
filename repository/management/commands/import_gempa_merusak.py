import datetime
import openpyxl
from django.core.management.base import BaseCommand, CommandError
from repository.models import GempaMemusak

BULAN_MAP = {
    'januari': 1, 'februari': 2, 'maret': 3, 'april': 4,
    'mei': 5, 'juni': 6, 'juli': 7, 'agustus': 8,
    'september': 9, 'oktober': 10, 'november': 11, 'desember': 12,
    'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4,
    'jun': 6, 'jul': 7, 'agt': 8, 'ags': 8,
    'sep': 9, 'okt': 10, 'nov': 11, 'des': 12,
}


def parse_tanggal(text):
    if not text:
        return None
    text = str(text).strip()
    # Remove any newline artifacts (e.g. "12 Agustus 2025\n2025")
    text = text.split('\n')[0].strip()
    parts = text.split()
    if len(parts) != 3:
        return None
    try:
        day = int(parts[0])
        month = BULAN_MAP.get(parts[1].lower())
        year = int(parts[2])
        if not month:
            return None
        return datetime.date(year, month, day)
    except (ValueError, TypeError):
        return None


def safe_float(val):
    if val is None or str(val).strip() in ('-', ''):
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


class Command(BaseCommand):
    help = 'Import Gempa Merusak data from xlsx file'

    def add_arguments(self, parser):
        parser.add_argument('xlsx_path', type=str, help='Path to the xlsx file')
        parser.add_argument('--clear', action='store_true', help='Clear existing records before import')

    def handle(self, *args, **options):
        path = options['xlsx_path']
        try:
            wb = openpyxl.load_workbook(path)
        except Exception as e:
            raise CommandError(f'Cannot open file: {e}')

        ws = wb.active

        if options['clear']:
            count, _ = GempaMemusak.objects.all().delete()
            self.stdout.write(f'Cleared {count} existing records.')

        events = []
        current = None

        for r in range(3, ws.max_row + 1):
            no_val = ws.cell(r, 1).value
            col2   = ws.cell(r, 2).value
            col3   = ws.cell(r, 3).value
            col9   = ws.cell(r, 9).value
            col10  = ws.cell(r, 10).value

            is_main = isinstance(no_val, int) or (
                isinstance(no_val, str) and str(no_val).strip().isdigit()
            )

            if is_main:
                if current:
                    events.append(current)
                ot_raw = col3
                ot = ot_raw if isinstance(ot_raw, datetime.time) else None

                current = {
                    'no':         int(no_val),
                    'tanggal_text': str(col2).strip() if col2 else '',
                    'origin_time': ot,
                    'lat':        safe_float(ws.cell(r, 4).value),
                    'lon':        safe_float(ws.cell(r, 5).value),
                    'depth':      safe_float(ws.cell(r, 6).value),
                    'mag':        safe_float(ws.cell(r, 7).value),
                    'lokasi':     ws.cell(r, 8).value or '',
                    'sumber':     ws.cell(r, 11).value or '',
                    'wilayah_parts': [],
                    'wilayah_merasakan_parts': [str(col9)] if col9 else [],
                    'korban_parts': [str(col10)] if col10 else [],
                    'tsunami': None,
                }
            elif current and col2 is not None:
                col2_str = str(col2).strip()
                if col2_str == 'Tsunami':
                    current['tsunami'] = True
                elif col2_str == 'Tidak Tsunami':
                    current['tsunami'] = False
                elif col2_str:
                    current['wilayah_parts'].append(col2_str)

                if col9:
                    current['wilayah_merasakan_parts'].append(str(col9))
                if col10:
                    current['korban_parts'].append(str(col10))
            elif current and col9:
                current['wilayah_merasakan_parts'].append(str(col9))
                if col10:
                    current['korban_parts'].append(str(col10))

        if current:
            events.append(current)

        created_count = 0
        updated_count = 0

        for e in events:
            tanggal = parse_tanggal(e['tanggal_text'])
            lokasi_val = str(e['lokasi']).strip()
            if lokasi_val not in ('Darat', 'Laut'):
                lokasi_val = ''

            wilayah = ', '.join(p for p in e['wilayah_parts'] if p)
            wilayah_merasakan = '\n'.join(p for p in e['wilayah_merasakan_parts'] if p)
            korban_kerusakan = '\n'.join(p for p in e['korban_parts'] if p)

            obj, created = GempaMemusak.objects.update_or_create(
                no=e['no'],
                defaults=dict(
                    tanggal_text      = e['tanggal_text'],
                    tanggal           = tanggal,
                    wilayah           = wilayah,
                    origin_time       = e['origin_time'],
                    latitude          = e['lat'],
                    longitude         = e['lon'],
                    depth_km          = e['depth'],
                    magnitude         = e['mag'],
                    lokasi            = lokasi_val,
                    tsunami           = e['tsunami'],
                    wilayah_merasakan = wilayah_merasakan,
                    korban_kerusakan  = korban_kerusakan,
                    sumber            = str(e['sumber']).strip(),
                ),
            )
            if created:
                created_count += 1
            else:
                updated_count += 1

        self.stdout.write(self.style.SUCCESS(
            f'Done. Created: {created_count}, Updated: {updated_count}, Total events: {len(events)}'
        ))
