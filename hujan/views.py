# hujan/views.py

from django.shortcuts import render, redirect, get_object_or_404 # Add redirect and get_object_or_404
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.http import HttpResponse
from .models import Hujan
from django.db.models import Max, Count, Q, Sum
from django.utils import timezone
from datetime import datetime
import json
from .forms import HujanForm, KATEGORI_CHOICES
from jadwal.models import Pegawai


def _filter_hujan(request):
    """Apply the daftar_hujan filters (start, end, kategori, petugas, q) from the
    request query string. Returns (queryset, filters_dict) so the list view and the
    Excel export stay in sync."""
    qs = Hujan.objects.all().order_by('-tanggal')
    f = {k: request.GET.get(k, '').strip()
         for k in ('start', 'end', 'kategori', 'petugas', 'q')}
    if f['start']:
        qs = qs.filter(tanggal__gte=f['start'])
    if f['end']:
        qs = qs.filter(tanggal__lte=f['end'])
    if f['kategori']:
        qs = qs.filter(kategori=f['kategori'])
    if f['petugas']:
        qs = qs.filter(petugas=f['petugas'])
    if f['q']:
        qs = qs.filter(keterangan__icontains=f['q'])
    return qs, f


def daftar_hujan(request):
    qs, f = _filter_hujan(request)
    f_start, f_end = f['start'], f['end']
    f_kategori, f_petugas, f_q = f['kategori'], f['petugas'], f['q']

    paginator = Paginator(qs, 10)
    page = request.GET.get('page')
    try:
        hujan_records = paginator.page(page)
    except PageNotAnInteger:
        hujan_records = paginator.page(1)
    except EmptyPage:
        hujan_records = paginator.page(paginator.num_pages)

    # Preserve active filters across pagination links (drop the page param).
    params = request.GET.copy()
    params.pop('page', None)
    querystring = params.urlencode()

    petugas_list = (Pegawai.objects.filter(tanggal_keluar__isnull=True)
                    .order_by('nama').values_list('nama', flat=True))

    context = {
        'hujan_records': hujan_records,
        'kategori_choices': KATEGORI_CHOICES,
        'petugas_list': petugas_list,
        'f_start': f_start,
        'f_end': f_end,
        'f_kategori': f_kategori,
        'f_petugas': f_petugas,
        'f_q': f_q,
        'querystring': querystring,
        'has_filters': any([f_start, f_end, f_kategori, f_petugas, f_q]),
    }
    return render(request, 'hujan/daftar_hujan.html', context)

def query_laporan_hujan(request):
    start_date = request.GET.get('start')
    end_date = request.GET.get('end')
    
    hujan_records = []
    stats = {}
    
    # Data untuk Chart (List kosong default)
    chart_dates = []
    chart_obs = []
    pie_labels = []
    pie_data = []

    if start_date and end_date:
        # 1. Filter Data (Urutkan tanggal asc untuk grafik)
        hujan_records = Hujan.objects.filter(
            tanggal__range=[start_date, end_date]
        ).order_by('tanggal')
        total_hujan = hujan_records.aggregate(Sum('obs'))['obs__sum'] or 0

        if hujan_records.exists():
            # --- LOGIKA NARASI (Hanya Obs) ---
            # Cari Max Obs
            max_hujan = hujan_records.aggregate(Max('obs'))
            max_val = max_hujan['obs__max']
            
            # Cari object detail dari nilai max tersebut
            max_obj = hujan_records.filter(obs=max_val).first()
            
            # Hitung Hari Hujan (Obs > 0)
            total_hari_hujan = hujan_records.filter(obs__gt=0).count()
            
            # Nama Bulan & Tahun untuk Narasi
            dt_obj = datetime.strptime(start_date, '%Y-%m-%d')
            bulan_tahun = dt_obj.strftime('%B %Y') # Contoh: November 2025

            stats = {
                'bulan_str': bulan_tahun,
                'total_hari_hujan': total_hari_hujan,
                'max_val': max_val,
                'max_date': max_obj.tanggal if max_obj else None,
                'max_kategori': max_obj.kategori if max_obj else '-',
                'total_hujan': total_hujan,
            }

            # --- PERSIAPAN DATA BAR CHART (OBS) ---
            for h in hujan_records:
                # Format tanggal jadi angka tanggal saja (1, 2, 3...) agar sumbu X rapi
                chart_dates.append(h.tanggal.day) 
                chart_obs.append(h.obs)

            # --- PERSIAPAN DATA PIE CHART (GROUP BY KATEGORI) ---
            # Menghitung jumlah kemunculan setiap kategori
            kategori_stats = hujan_records.values('kategori').annotate(total=Count('kategori')).order_by('-total')
            
            for item in kategori_stats:
                pie_labels.append(item['kategori']) # e.g., "Sedang", "Nihil"
                pie_data.append(item['total'])      # e.g., 5, 10

    context = {
        'hujan_records': hujan_records,
        'start_date': start_date,
        'end_date': end_date,
        'stats': stats,
        # Dump data ke JSON string untuk JS
        'chart_dates': json.dumps(chart_dates),
        'chart_obs': json.dumps(chart_obs),
        'pie_labels': json.dumps(pie_labels),
        'pie_data': json.dumps(pie_data),
    }

    return render(request, 'hujan/laporan_query.html', context)

def tambah_hujan(request):
    if request.method == 'POST':
        form = HujanForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('daftar_hujan')
    else:
        form = HujanForm()

    context = {
        'form': form,
        'title': 'Tambah Data Hujan',
        'submit_label': 'Simpan',
    }
    return render(request, 'hujan/form_hujan.html', context)

def edit_hujan(request, id):
    # Get the specific record or return 404 if not found
    hujan_instance = get_object_or_404(Hujan, id=id)

    if request.method == 'POST':
        form = HujanForm(request.POST, instance=hujan_instance)
        if form.is_valid():
            form.save()
            return redirect('daftar_hujan') # Redirect back to list after saving
    else:
        # Pre-fill the form with existing data
        form = HujanForm(instance=hujan_instance)

    context = {
        'form': form,
        'title': 'Edit Data Hujan',
        'submit_label': 'Simpan Perubahan',
    }
    return render(request, 'hujan/form_hujan.html', context)


def export_hujan_excel(request):
    """Export the (filtered) hujan list to .xlsx. Uses the same filters as
    daftar_hujan, so the download matches what's shown in the table."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    qs, _f = _filter_hujan(request)

    wb = Workbook()
    ws = wb.active
    ws.title = "Hujan"

    headers = ['Tanggal', 'Obs (mm)', 'Hilman', 'Kategori', 'Keterangan', 'Petugas']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for h in qs:
        ws.append([
            h.tanggal.strftime('%Y-%m-%d') if h.tanggal else '',
            h.obs,
            h.hilman,
            h.kategori,
            h.keterangan or '',
            h.petugas,
        ])

    for i, width in enumerate([14, 10, 10, 16, 40, 18], start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    filename = "hujan_%s.xlsx" % timezone.now().strftime('%Y%m%d_%H%M%S')
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="%s"' % filename
    wb.save(response)
    return response